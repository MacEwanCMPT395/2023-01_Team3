from PyQt5.QtWidgets import *
import sys

import openpyxl
# If you are using pycharm you can install openpyxl by right-clicking the red squiggles and check its
# context actions, otherwise: pip install openpyxl
# https://stackoverflow.com/questions/38364404/how-to-install-openpyxl-with-pip


# This is a very basic GUI for purposes of displaying the xlsx file data and getting input. It is very work in
# progress. So I am still working out how pyqt5 gui works, some elements are not working like how I expect them to,
# especially with the HBoxlayouts and VBoxlayouts. There are design issues that I will keep working on to make it
# look better, but the GUI can allow the user to open a xlsx file from anywhere on their computer to be read. I
# wanted to make sure that the file can be opened and the data can be read in. I also made a section for adding
# inputs (just the year so far) and a button that will display what was entered to ensure that the inputs can also be
# read in for calculations on what day classes start..etc. Ran into issues with my laptop at work, and can't run qt
# designer until I am able to fix my laptop. But the methods I made should still transfer over to designer if someone
# decides coding the GUI by "hand" is too difficult. The designer will take care of building the look, just need to
# write a few lines of code to make certain elements interactive like how I did with my buttons.

# Note: If anyone is going to try to make a different GUI look in designer, after you finish making it, the program will
# save it as a .ui file. To convert it to a .py file all you need to do is open your terminal, go to the directory
# where the file is saved and type: pyuic5 -x <filename>.ui -o <filename>.py

# https://www.youtube.com/watch?v=5K__zwBj_nY 7:00 min is where the terminal command is mentioned
# Another method to get it working that is not using the command line. Need to import: from PyQt5 import uic
# https://www.youtube.com/watch?v=MOItX2aKTGc 13:30 min is where this method is mentioned.

class Main(QMainWindow):

    def __init__(self):
        super(Main, self).__init__()
        self.setWindowTitle("Scheduler")

        # This is the main layout for the GUI (HBox layout lays everything in it horizontally from each other)
        main_layout = QHBoxLayout()
        main_layout.setSpacing(100)

        # HBoxlayout to hold the interactive "panel" on the left side of gui and the tables to display on the right
        h_layout = QHBoxLayout()
        # VBoxlayout to group together the interactive elements of the GUI on the left side but vertically
        v_layout = QVBoxLayout()
        v_layout.setSpacing(5)
        # self.setLayout(main_layout)

        # Grouped together labels and their buttons or input fields
        file_label = QLabel("File")
        file_button = QPushButton("Open File")
        file_label.setMaximumSize(100, 50)
        file_button.setMaximumSize(100, 50)

        # Textbox inputs from user
        # Displaying textbox/lineedit inputs as a dev check to make sure we can get input for calculations
        button_show_input = QPushButton("Display inputs")
        self.label_year = QLabel("Year")
        self.textbox_year = QLineEdit()
        button_show_input.setMaximumSize(100, 50)
        self.label_year.setMaximumSize(100, 50)
        self.textbox_year.setMaximumSize(100, 50)

        # Adding interactive widgets to the VBoxlayout
        v_layout.addWidget(file_label)
        v_layout.addWidget(file_button)

        v_layout.addWidget(self.label_year)
        v_layout.addWidget(self.textbox_year)
        v_layout.addWidget(button_show_input)

        # This is the tableWidget to display data from xlsx file; unsure what our data is going to look like
        # Can modify later
        self.table_widget = QTableWidget()
        self.table_widget.setGeometry(100, 100, 100, 100)

        # Adding the VBoxlayout with the interactive elements to the HBoxlayout
        h_layout.addLayout(v_layout)
        h_layout.addWidget(self.table_widget)

        # Had to make a widget and set its layout to the main_layout with all the elements to display anything
        # Will try to figure out why that is the ca
        main_layout.addLayout(h_layout)
        widget = QWidget()
        widget.setLayout(main_layout)
        self.setCentralWidget(widget)

        # Connecting the buttons to their functions (what happens when they are clicked)
        file_button.clicked.connect(self.open_file)
        button_show_input.clicked.connect(self.usr_inputs)

    # This function will open an explorer window and allows the user to select the xslx file to open
    # this is connected to the button "file_button" - action event
    def open_file(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "",
                                                   "CSV Files (*.csv);;All Files (*)", options=options)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        self.table_widget.setRowCount(sheet.max_row)
        self.table_widget.setColumnCount(sheet.max_column)

        list_values = list(sheet.values)
        # testing code that prints xlsx file data into the console
        # for value in values:
        #     print(value)
        self.table_widget.setHorizontalHeaderLabels(list_values[0])

        # loop to max rows and each row loops over each column
        # currently unsure how to the data file will look
        # just looping through the file to display data
        # can refactor it later depending on how the data we get looks
        row_index = 0
        for value_tuple in list_values[1:]:
            col_index = 0
            # print(value_tuple) # dev check
            for value in value_tuple:
                self.table_widget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1

    # Temp function as a dev check to display user inputs
    def usr_inputs(self):
        if self.textbox_year.text() != "":
            input_parameters = QMessageBox()
            input_parameters.setText(self.textbox_year.text())
            input_parameters.exec()


def main():
    app = QApplication(sys.argv)
    window = Main()
    window.showMaximized()
    app.exec_()


if __name__ == '__main__':
    main()
